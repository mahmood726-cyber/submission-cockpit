from __future__ import annotations

import csv
import importlib.util
import json
import re
import shutil
import subprocess
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path


OPENPYXL_AVAILABLE = importlib.util.find_spec("openpyxl") is not None

if OPENPYXL_AVAILABLE:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill


PROJECT_ROOT = Path(__file__).resolve().parents[1]
USER_ROOT = PROJECT_ROOT.parents[0]
DATA_SOURCE = PROJECT_ROOT / "data-source"
EXPORTS_DIR = PROJECT_ROOT / "exports"
MANIFEST_DIR = EXPORTS_DIR / "project-manifests"
TRACKER_XLSX = DATA_SOURCE / "rewrite-workbook.xlsx"
TRACKER_CSV = DATA_SOURCE / "rewrite-tracker.csv"
TRACKER_JSON = DATA_SOURCE / "rewrite-tracker.json"
COCKPIT_JSON = PROJECT_ROOT / "submission-cockpit.json"
DATA_JSON = PROJECT_ROOT / "data.json"
DATA_JS = PROJECT_ROOT / "data.js"
E156_PAPER_JSON = PROJECT_ROOT / "e156-submission" / "paper.json"
E156_INDEX_HTML = PROJECT_ROOT / "e156-submission" / "index.html"
E156_BUILDER = USER_ROOT / "E156-framework" / "scripts" / "build_e156_bundle.py"
EXTERNAL_E156_ROOT = Path("/mnt/c/E156")
EXTERNAL_WORKBOOK_SOURCE = EXTERNAL_E156_ROOT / "rewrite-workbook.txt"
EXTERNAL_DAILY_SYNC = EXTERNAL_E156_ROOT / "scripts" / "daily_sync.bat"


SOURCE_FILES = {
    "portfolio": USER_ROOT / "ResearchConstellation" / "portfolio-data.json",
    "ops": USER_ROOT / "PortfolioOps" / "ops-readiness.json",
    "citation": USER_ROOT / "CitationWorkbench" / "citation-readiness.json",
    "authorship": USER_ROOT / "AuthorshipLedger" / "authorship-ledger.json",
    "deposit": USER_ROOT / "AuthorshipLedger" / "deposit-manifest.json",
    "external_rewrite_workbook": EXTERNAL_WORKBOOK_SOURCE,
}

SNAPSHOT_FILES = {
    "portfolio": DATA_SOURCE / "portfolio-data.snapshot.json",
    "ops": DATA_SOURCE / "ops-readiness.snapshot.json",
    "citation": DATA_SOURCE / "citation-readiness.snapshot.json",
    "authorship": DATA_SOURCE / "authorship-ledger.snapshot.json",
    "deposit": DATA_SOURCE / "deposit-manifest.snapshot.json",
    "external_rewrite_workbook": DATA_SOURCE / "rewrite-workbook.snapshot.txt",
    "external_audit": DATA_SOURCE / "e156-audit-report.snapshot.json",
    "external_maintenance": DATA_SOURCE / "e156-maintenance-report.snapshot.json",
    "external_verification": DATA_SOURCE / "e156-verification-report.snapshot.json",
}


TRACKER_COLUMNS = [
    "project_id",
    "project_name",
    "slug",
    "project_path",
    "tier",
    "resolved_status",
    "target_journal",
    "portfolio_priority",
    "ops_readiness_score",
    "citation_readiness_score",
    "workflow_readiness_score",
    "recommended_license",
    "source_paper_signal",
    "source_protocol_signal",
    "existing_pages_url",
    "deposit_state",
    "rewrite_workbook_source",
    "rewrite_present",
    "rewrite_valid",
    "rewrite_word_count",
    "rewrite_sentence_count",
    "rewrite_matches_current",
    "rewrite_status",
    "publish_now",
    "e156_body_status",
    "protocol_status",
    "dashboard_status",
    "code_release_status",
    "mit_license_status",
    "github_status",
    "pages_status",
    "pdf_status",
    "synthesis_status",
    "source_article_path",
    "github_repo_url",
    "github_pages_url",
    "paper_pdf_path",
    "synthesis_article_url",
    "notes",
    "last_updated",
]

MANUAL_COLUMNS = {
    "rewrite_status",
    "publish_now",
    "e156_body_status",
    "protocol_status",
    "dashboard_status",
    "code_release_status",
    "mit_license_status",
    "github_status",
    "pages_status",
    "pdf_status",
    "synthesis_status",
    "source_article_path",
    "github_repo_url",
    "github_pages_url",
    "paper_pdf_path",
    "synthesis_article_url",
    "notes",
    "last_updated",
}

STATUS_CHOICES = {
    "rewrite_status": {"not-started", "rewriting", "rewritten", "approved", "published"},
    "publish_now": {"yes", "no"},
    "e156_body_status": {"missing", "draft", "ready", "published"},
    "protocol_status": {"missing", "draft", "ready", "published"},
    "dashboard_status": {"missing", "draft", "ready", "published"},
    "code_release_status": {"missing", "draft", "ready", "published"},
    "mit_license_status": {"missing", "planned", "ready"},
    "github_status": {"missing", "draft", "ready", "published"},
    "pages_status": {"missing", "draft", "ready", "published"},
    "pdf_status": {"not-needed", "pending", "ready", "uploaded"},
    "synthesis_status": {"not-configured", "blocked", "ready", "uploaded"},
}

READY_STATUSES = {"ready", "published"}
REWRITE_READY_STATUSES = {"rewritten", "approved", "published"}
PUBLISHED_STATUSES = {"published"}
PRIORITY_ORDER = {"high": 0, "medium": 1, "low": 2}
WORKBOOK_BOUNDARY_WORDS = [
    "limit",
    "cannot",
    "may not",
    "unclear",
    "uncertain",
    "caution",
    "scope",
    "boundary",
    "harm",
    "constrain",
    "generali",
    "restrict",
    "warrant",
    "however",
    "not ",
    "does not",
    "only",
    "unable",
    "lack",
    "miss",
    "exclude",
    "narrow",
    "depend",
    "assume",
    "imprecis",
    "approximate",
    "not yet",
    "not address",
    "not extend",
    "not capture",
    "not replace",
    "not detect",
    "not account",
    "not valid",
    "not includ",
    "not support",
    "not scale",
    "not assess",
    "not verify",
    "remains",
    "rather than",
    "non-exclusive",
    "should be interpret",
    "caveat",
]


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def ensure_dirs() -> None:
    DATA_SOURCE.mkdir(parents=True, exist_ok=True)
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    MANIFEST_DIR.mkdir(parents=True, exist_ok=True)


def load_json(path: Path) -> dict[str, object]:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: dict[str, object] | list[object]) -> None:
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def sync_snapshots() -> None:
    for key, source_path in SOURCE_FILES.items():
        if not source_path.exists():
            raise FileNotFoundError(f"Required upstream file is missing: {source_path}")
        shutil.copy2(source_path, SNAPSHOT_FILES[key])

    optional_sources = {
        "external_audit": EXTERNAL_E156_ROOT / "audit-report.json",
        "external_maintenance": EXTERNAL_E156_ROOT / "maintenance-report.json",
        "external_verification": EXTERNAL_E156_ROOT / "verification-report.json",
    }
    for key, source_path in optional_sources.items():
        if source_path.exists():
            shutil.copy2(source_path, SNAPSHOT_FILES[key])


def slug_or_id(row: dict[str, str]) -> str:
    return (row.get("project_id") or row.get("slug") or "").strip()


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_choice(column: str, value: object) -> str:
    text = normalize_text(value).lower()
    if column not in STATUS_CHOICES:
        return normalize_text(value)
    if text in STATUS_CHOICES[column]:
        return text
    defaults = {
        "rewrite_status": "not-started",
        "publish_now": "no",
        "e156_body_status": "missing",
        "protocol_status": "missing",
        "dashboard_status": "missing",
        "code_release_status": "missing",
        "mit_license_status": "missing",
        "github_status": "missing",
        "pages_status": "missing",
        "pdf_status": "not-needed",
        "synthesis_status": "not-configured",
    }
    return defaults[column]


def normalize_windows_path(value: object) -> str:
    text = re.sub(r"\\+", r"\\", normalize_text(value).replace("/", "\\"))
    return text.rstrip("\\").lower()


def normalize_name_key(value: object) -> str:
    text = normalize_text(value).lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def path_leaf(value: object) -> str:
    normalized = normalize_windows_path(value)
    if not normalized:
        return ""
    parts = [part for part in normalized.split("\\") if part and not part.endswith(":")]
    return parts[-1] if parts else ""


def windows_to_local_path(value: object) -> Path | None:
    text = normalize_text(value).replace("/", "\\")
    match = re.match(r"^([A-Za-z]):\\(.*)$", text)
    if not match:
        return None
    drive = match.group(1).lower()
    remainder = match.group(2).replace("\\", "/")
    return Path(f"/mnt/{drive}/{remainder}")


def local_to_windows_path(path: Path) -> str:
    text = str(path)
    if text == "/mnt/c":
        return r"C:\\"
    if text.startswith("/mnt/c/"):
        return r"C:" + "\\" + text[len("/mnt/c/"):].replace("/", "\\")
    return text.replace("/", "\\")


def slugify_text(value: object) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-") or "workbook-entry"


def tracker_base_id(value: object) -> str:
    return normalize_text(value).split("::", 1)[0]


def build_items_by_id(items: list[dict[str, object]]) -> dict[str, list[dict[str, object]]]:
    grouped: dict[str, list[dict[str, object]]] = {}
    for item in items:
        grouped.setdefault(normalize_text(item.get("id")), []).append(item)
    return grouped


def matched_item(
    items_by_id: dict[str, list[dict[str, object]]],
    raw_id: str,
    *,
    project_name: object,
    project_path: object,
    project_slug: object = "",
) -> dict[str, object]:
    candidates = items_by_id.get(normalize_text(raw_id), [])
    if not candidates:
        return {}
    if len(candidates) == 1:
        return candidates[0]

    name_key = normalize_name_key(project_name)
    path_key = normalize_windows_path(project_path)
    path_tail = path_leaf(project_path)
    slug_key = normalize_name_key(project_slug)

    def score(item: dict[str, object]) -> tuple[int, str]:
        item_name_key = normalize_name_key(item.get("name"))
        item_path_key = normalize_windows_path(item.get("path"))
        item_path_tail = path_leaf(item.get("path"))
        item_slug_key = normalize_name_key(item.get("slug"))
        item_record_key = normalize_name_key(item.get("recordUrl"))

        points = 0
        if name_key and item_name_key == name_key:
            points += 100
        if path_key and item_path_key and item_path_key == path_key:
            points += 80
        if path_tail and item_path_tail and item_path_tail == path_tail:
            points += 20
        if slug_key and item_slug_key and item_slug_key == slug_key:
            points += 40
        if name_key and item_slug_key and name_key in item_slug_key:
            points += 25
        if name_key and item_record_key and name_key in item_record_key:
            points += 10
        return points, normalize_text(item.get("name"))

    return max(candidates, key=score)


def split_workbook_sentences(text: str) -> list[str]:
    return [part.strip() for part in re.split(r"(?<=[.!?])\s+(?=[A-Z0-9])", text.strip()) if part.strip()]


def validate_workbook_rewrite(body: str) -> tuple[bool, list[str], int, int]:
    issues: list[str] = []
    words = len(body.split())
    sentences = split_workbook_sentences(body)

    if words > 156:
        issues.append(f"Word count {words} exceeds 156")
    if words < 100:
        issues.append(f"Word count {words} is too short")
    if len(sentences) != 7:
        issues.append(f"Sentence count {len(sentences)} is not 7")
    if "\n\n" in body:
        issues.append("Body contains blank lines")
    if re.search(r"^\s*#", body, re.MULTILINE):
        issues.append("Body contains markdown headings")
    if re.search(r"https?://", body):
        issues.append("Body contains URLs")
    if len(sentences) >= 5 and not any(re.search(r"\d", sentences[index]) for index in range(2, 5)):
        issues.append("Sentences 3 to 5 do not contain a quantitative result")
    if len(sentences) >= 7 and not any(word in sentences[6].lower() for word in WORKBOOK_BOUNDARY_WORDS):
        issues.append("Sentence 7 lacks limitation language")

    return not issues, issues, words, len(sentences)


def parse_external_rewrite_workbook(path: Path) -> list[dict[str, object]]:
    text = path.read_text(encoding="utf-8")
    entries: list[dict[str, object]] = []
    blocks = re.split(r"^={50,}$", text, flags=re.MULTILINE)

    for block in blocks:
        header_match = re.search(r"\[(\d+)/\d+\]\s+(.+)", block)
        if not header_match:
            continue

        path_match = re.search(r"^PATH:\s*(.+)$", block, re.MULTILINE)
        if not path_match:
            continue

        title_match = re.search(r"^TITLE:\s*(.+)$", block, re.MULTILINE)
        type_match = re.search(r"^TYPE:\s*(.+?)\s+\|\s+ESTIMAND:\s*(.+)$", block, re.MULTILINE)
        data_match = re.search(r"^DATA:\s*(.+)$", block, re.MULTILINE)

        current_match = re.search(r"CURRENT BODY \(\d+ words\):\n(.+?)(?=\nYOUR REWRITE)", block, re.DOTALL)
        rewrite_match = re.search(r"YOUR REWRITE \(at most 156 words, 7 sentences\):\n(.+?)$", block, re.DOTALL)

        current = current_match.group(1).strip() if current_match else ""
        rewrite = rewrite_match.group(1).strip() if rewrite_match else ""
        rewrite_matches_current = bool(rewrite) and rewrite == current
        rewrite_present = bool(rewrite) and not rewrite_matches_current
        rewrite_valid, issues, word_count, sentence_count = validate_workbook_rewrite(rewrite) if rewrite_present else (False, [], 0, 0)

        entries.append(
            {
                "index": int(header_match.group(1)),
                "name": header_match.group(2).strip(),
                "title": title_match.group(1).strip() if title_match else header_match.group(2).strip(),
                "type": type_match.group(1).strip() if type_match else "",
                "estimand": type_match.group(2).strip() if type_match else "",
                "data": data_match.group(1).strip() if data_match else "",
                "path": path_match.group(1).strip(),
                "normalizedPath": normalize_windows_path(path_match.group(1)),
                "current": current,
                "rewrite": rewrite,
                "rewritePresent": rewrite_present,
                "rewriteValid": rewrite_valid,
                "rewriteIssues": issues,
                "rewriteWordCount": word_count,
                "rewriteSentenceCount": sentence_count,
                "rewriteMatchesCurrent": rewrite_matches_current,
            }
        )

    return entries


def load_repo_mapping() -> dict[str, str]:
    path = EXTERNAL_E156_ROOT / "scripts" / "repo_mapping.json"
    if not path.exists():
        return {}
    raw = load_json(path)
    return {normalize_windows_path(key): normalize_text(value) for key, value in raw.items()}


def load_submission_metadata(path_str: str) -> dict[str, str]:
    local_path = windows_to_local_path(path_str)
    defaults = {
        "title": "",
        "slug": "",
        "type": "",
        "estimand": "",
        "data": "",
        "body": "",
        "bodyValid": "no",
        "repoUrl": "",
        "siteUrl": "",
        "sourceArticlePath": "",
    }
    if not local_path:
        return defaults

    config_path = local_path / "e156-submission" / "config.json"
    paper_path = local_path / "e156-submission" / "paper.json"
    paper_md_path = local_path / "e156-submission" / "paper.md"

    config: dict[str, object] = {}
    paper: dict[str, object] = {}
    if config_path.exists():
        try:
            config = load_json(config_path)
        except (OSError, json.JSONDecodeError):
            config = {}
    if paper_path.exists():
        try:
            paper = load_json(paper_path)
        except (OSError, json.JSONDecodeError):
            paper = {}

    notes = config.get("notes")
    data_note = ""
    if isinstance(notes, dict):
        data_note = normalize_text(notes.get("data"))

    body = normalize_text(paper.get("body")) or normalize_text(config.get("body"))
    body_valid = False
    if body:
        body_valid, _, _, _ = validate_workbook_rewrite(body)

    if paper_path.exists():
        source_article_path = local_to_windows_path(paper_path)
    elif paper_md_path.exists():
        source_article_path = local_to_windows_path(paper_md_path)
    else:
        source_article_path = ""

    return {
        "title": normalize_text(paper.get("title")) or normalize_text(config.get("title")) or local_path.name,
        "slug": normalize_text(paper.get("slug")) or normalize_text(config.get("slug")) or slugify_text(local_path.name),
        "type": normalize_text(paper.get("type")) or normalize_text(config.get("type")) or "methods",
        "estimand": normalize_text(paper.get("primary_estimand")) or normalize_text(config.get("primary_estimand")),
        "data": data_note or normalize_text(paper.get("summary")) or normalize_text(config.get("summary")),
        "body": body,
        "bodyValid": "yes" if body_valid else "no",
        "repoUrl": normalize_text(config.get("repo_url")),
        "siteUrl": normalize_text(config.get("site_url")) or normalize_text(config.get("dashboard_url")),
        "sourceArticlePath": source_article_path,
    }


def external_slug_from_path(path_str: str, fallback_slug: str) -> str:
    normalized = normalize_windows_path(path_str)
    parts = [part for part in normalized.split("\\") if part and not part.endswith(":")]
    if not parts:
        return fallback_slug or "external-project"
    return slugify_text("-".join(parts)) or fallback_slug or "external-project"


def inspect_project_path(path_str: str) -> dict[str, object]:
    local_path = windows_to_local_path(path_str)
    if not local_path:
        return {
            "exists": False,
            "hasGit": False,
            "hasPaper": False,
            "hasProtocol": False,
            "hasDashboard": False,
            "hasPages": False,
            "hasMITLicense": False,
        }

    e156_dir = local_path / "e156-submission"
    root_license = local_path / "LICENSE"
    mit_license = False
    if root_license.exists():
        try:
            mit_license = "mit license" in root_license.read_text(encoding="utf-8", errors="ignore").lower()
        except OSError:
            mit_license = False

    has_git = (local_path / ".git").exists()
    if not has_git:
        try:
            probe = subprocess.run(
                ["git", "-C", str(local_path), "rev-parse", "--is-inside-work-tree"],
                capture_output=True,
                text=True,
                check=False,
                timeout=5,
            )
            has_git = probe.returncode == 0 and probe.stdout.strip() == "true"
        except (OSError, subprocess.SubprocessError):
            has_git = False

    return {
        "exists": local_path.exists(),
        "hasGit": has_git,
        "hasPaper": (e156_dir / "paper.md").exists() or (e156_dir / "paper.json").exists(),
        "hasProtocol": (e156_dir / "protocol.md").exists(),
        "hasDashboard": (local_path / "index.html").exists() or (e156_dir / "assets" / "dashboard.html").exists(),
        "hasPages": (local_path / "index.html").exists() or (local_path / ".nojekyll").exists(),
        "hasMITLicense": mit_license,
    }


def priority_for_workbook_entry(entry: dict[str, object], path_state: dict[str, object]) -> str:
    if entry["rewriteValid"]:
        return "high"
    if entry["rewritePresent"] or path_state["exists"]:
        return "medium"
    return "low"


def load_optional_json(path: Path) -> dict[str, object]:
    if not path.exists():
        return {}
    return load_json(path)


def build_external_e156_status(workbook_entries: list[dict[str, object]]) -> dict[str, object]:
    audit_report = load_optional_json(SNAPSHOT_FILES["external_audit"])
    maintenance_report = load_optional_json(SNAPSHOT_FILES["external_maintenance"])
    verification_report = load_optional_json(SNAPSHOT_FILES["external_verification"])

    matched_entries = [entry for entry in workbook_entries if entry["rewritePresent"]]
    valid_entries = [entry for entry in matched_entries if entry["rewriteValid"]]

    return {
        "root": r"C:\E156",
        "workbookSource": r"C:\E156\rewrite-workbook.txt",
        "dailySyncPath": r"C:\E156\scripts\daily_sync.bat" if EXTERNAL_DAILY_SYNC.exists() else "",
        "workbookProjectCount": len(workbook_entries),
        "workbookRewriteCount": len(matched_entries),
        "workbookValidRewriteCount": len(valid_entries),
        "auditSummary": {
            "p0": audit_report.get("p0_count"),
            "p1": audit_report.get("p1_count"),
            "p2": audit_report.get("p2_count"),
            "cleanCount": audit_report.get("clean_count"),
            "totalProjects": audit_report.get("total_projects"),
        } if audit_report else {},
        "maintenanceOverallOk": maintenance_report.get("overall_ok") if maintenance_report else None,
        "verificationOverallOk": verification_report.get("overall_ok") if verification_report else None,
    }


def priority_for_project(citation: dict[str, object], ops: dict[str, object]) -> str:
    resolved = normalize_text(citation.get("resolvedStatus"))
    readiness = int(ops.get("readinessScore", 0))
    if resolved == "Submission ready" or readiness >= 85:
        return "high"
    if resolved in {"Active", "Review", "Accepted"} or readiness >= 70:
        return "medium"
    return "low"


def publish_default(citation: dict[str, object]) -> str:
    return "yes" if normalize_text(citation.get("resolvedStatus")) in {"Submission ready", "Active", "Review", "Accepted"} else "no"


def unique_indexed_project_id(
    project: dict[str, object],
    citation: dict[str, object],
    duplicate_counts: dict[str, int],
) -> str:
    raw_id = normalize_text(project.get("id"))
    if duplicate_counts.get(raw_id, 0) <= 1:
        return raw_id

    suffix = normalize_text(citation.get("slug")) or external_slug_from_path(
        normalize_text(project.get("path")),
        slugify_text(project.get("name")),
    )
    return f"{raw_id}::{suffix}"


def default_tracker_row(
    project_id: str,
    project: dict[str, object],
    citation: dict[str, object],
    authorship: dict[str, object],
    deposit: dict[str, object],
    ops: dict[str, object],
) -> dict[str, str]:
    existing_pages = normalize_text(citation.get("recordUrl"))
    recommended_license = normalize_text(authorship.get("recommendedLicense")) or "MIT"
    return {
        "project_id": project_id,
        "project_name": normalize_text(project.get("name")),
        "slug": normalize_text(citation.get("slug")) or normalize_text(project.get("id")),
        "project_path": normalize_text(project.get("path")),
        "tier": normalize_text(project.get("tierShortName")),
        "resolved_status": normalize_text(citation.get("resolvedStatus")) or normalize_text(project.get("statusLabel")),
        "target_journal": normalize_text(citation.get("targetJournal")),
        "portfolio_priority": priority_for_project(citation, ops),
        "ops_readiness_score": str(ops.get("readinessScore", 0)),
        "citation_readiness_score": str(citation.get("citationReadinessScore", 0)),
        "workflow_readiness_score": str(authorship.get("workflowReadinessScore", 0)),
        "recommended_license": recommended_license or "MIT",
        "source_paper_signal": "yes" if bool(citation.get("hasPaper")) else "no",
        "source_protocol_signal": "yes" if bool(citation.get("hasProtocol")) else "no",
        "existing_pages_url": existing_pages,
        "deposit_state": normalize_text(deposit.get("depositState")) or normalize_text(authorship.get("depositState")),
        "rewrite_workbook_source": r"C:\E156\rewrite-workbook.txt",
        "rewrite_present": "no",
        "rewrite_valid": "no",
        "rewrite_word_count": "",
        "rewrite_sentence_count": "",
        "rewrite_matches_current": "no",
        "rewrite_status": "not-started",
        "publish_now": publish_default(citation),
        "e156_body_status": "draft" if bool(citation.get("hasPaper")) else "missing",
        "protocol_status": "draft" if bool(citation.get("hasProtocol")) else "missing",
        "dashboard_status": "draft" if existing_pages else "missing",
        "code_release_status": "ready" if bool(ops.get("codeSignal")) else "missing",
        "mit_license_status": "planned" if recommended_license == "MIT" else "missing",
        "github_status": "missing",
        "pages_status": "draft" if existing_pages else "missing",
        "pdf_status": "pending" if publish_default(citation) == "yes" else "not-needed",
        "synthesis_status": "not-configured",
        "source_article_path": "",
        "github_repo_url": "",
        "github_pages_url": "",
        "paper_pdf_path": "",
        "synthesis_article_url": "",
        "notes": "",
        "last_updated": "",
    }


def default_external_tracker_row(
    workbook_entry: dict[str, object],
    repo_mapping: dict[str, str],
) -> dict[str, str]:
    path_str = normalize_text(workbook_entry["path"])
    path_state = inspect_project_path(path_str)
    metadata = load_submission_metadata(path_str)
    unique_slug = external_slug_from_path(path_str, metadata["slug"] or slugify_text(workbook_entry["name"]))
    repo_url = metadata["repoUrl"] or repo_mapping.get(normalize_windows_path(path_str), "")
    pages_url = metadata["siteUrl"]
    body_ready = workbook_entry["rewriteValid"] or metadata["bodyValid"] == "yes"
    body_draft = path_state["hasPaper"] or bool(metadata["body"])

    publish_now = (
        body_ready
        or (
            body_ready
            and path_state["hasProtocol"]
            and path_state["hasDashboard"]
            and path_state["hasMITLicense"]
        )
    )

    ops_score = min(
        100,
        20
        + 20 * int(bool(path_state["hasPaper"]))
        + 20 * int(bool(path_state["hasProtocol"]))
        + 20 * int(bool(path_state["hasDashboard"]))
        + 20 * int(bool(path_state["hasGit"])),
    )
    citation_score = min(
        100,
        25
        + 20 * int(bool(workbook_entry["rewritePresent"]))
        + 25 * int(bool(workbook_entry["rewriteValid"]))
        + 15 * int(bool(metadata["title"]))
        + 15 * int(bool(metadata["estimand"])),
    )
    workflow_score = min(
        100,
        20
        + 20 * int(bool(path_state["hasPaper"]))
        + 20 * int(bool(path_state["hasProtocol"]))
        + 20 * int(bool(path_state["hasDashboard"]))
        + 20 * int(bool(path_state["hasMITLicense"])),
    )

    return {
        "project_id": f"ext-{unique_slug}",
        "project_name": normalize_text(workbook_entry["name"]),
        "slug": unique_slug,
        "project_path": path_str,
        "tier": "Workbook-only",
        "resolved_status": "External E156 repo",
        "target_journal": "E156",
        "portfolio_priority": priority_for_workbook_entry(workbook_entry, path_state),
        "ops_readiness_score": str(ops_score),
        "citation_readiness_score": str(citation_score),
        "workflow_readiness_score": str(workflow_score),
        "recommended_license": "MIT",
        "source_paper_signal": "yes" if path_state["hasPaper"] else "no",
        "source_protocol_signal": "yes" if path_state["hasProtocol"] else "no",
        "existing_pages_url": pages_url,
        "deposit_state": "Not indexed",
        "rewrite_workbook_source": r"C:\E156\rewrite-workbook.txt",
        "rewrite_present": "yes" if workbook_entry["rewritePresent"] else "no",
        "rewrite_valid": "yes" if workbook_entry["rewriteValid"] else "no",
        "rewrite_word_count": str(workbook_entry["rewriteWordCount"]) if workbook_entry["rewritePresent"] else "",
        "rewrite_sentence_count": str(workbook_entry["rewriteSentenceCount"]) if workbook_entry["rewritePresent"] else "",
        "rewrite_matches_current": "yes" if workbook_entry["rewriteMatchesCurrent"] else "no",
        "rewrite_status": "rewritten" if workbook_entry["rewriteValid"] else ("rewriting" if workbook_entry["rewritePresent"] else "not-started"),
        "publish_now": "yes" if publish_now else "no",
        "e156_body_status": "ready" if body_ready else ("draft" if body_draft else "missing"),
        "protocol_status": "ready" if path_state["hasProtocol"] else "missing",
        "dashboard_status": "ready" if path_state["hasDashboard"] else "missing",
        "code_release_status": "ready" if path_state["hasGit"] else ("draft" if path_state["exists"] else "missing"),
        "mit_license_status": "ready" if path_state["hasMITLicense"] else "planned",
        "github_status": "published" if repo_url else "missing",
        "pages_status": "published" if pages_url else "missing",
        "pdf_status": "pending" if publish_now else "not-needed",
        "synthesis_status": "not-configured",
        "source_article_path": metadata["sourceArticlePath"] or (r"C:\E156\rewrite-workbook.txt" if workbook_entry["rewritePresent"] else ""),
        "github_repo_url": repo_url,
        "github_pages_url": pages_url,
        "paper_pdf_path": "",
        "synthesis_article_url": "",
        "notes": "Workbook-only project discovered from the C drive scan.",
        "last_updated": "",
    }


def load_tracker_csv() -> list[dict[str, str]]:
    if not TRACKER_CSV.exists():
        return []
    with TRACKER_CSV.open("r", encoding="utf-8", newline="") as handle:
        return [{key: normalize_text(value) for key, value in row.items()} for row in csv.DictReader(handle)]


def load_tracker_xlsx() -> list[dict[str, str]]:
    if not OPENPYXL_AVAILABLE or not TRACKER_XLSX.exists():
        return []
    workbook = load_workbook(TRACKER_XLSX, data_only=True)
    if "Tracker" not in workbook.sheetnames:
        return []
    sheet = workbook["Tracker"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_text(cell) for cell in rows[0]]
    loaded: list[dict[str, str]] = []
    for values in rows[1:]:
        if not any(values):
            continue
        row = {
            header: normalize_text(value)
            for header, value in zip(headers, values)
            if header
        }
        if slug_or_id(row):
            loaded.append(row)
    return loaded


def external_source_path(metadata: dict[str, str], rewrite_present: bool) -> str:
    return metadata["sourceArticlePath"] or (r"C:\E156\rewrite-workbook.txt" if rewrite_present else "")


def merge_tracker_rows(
    defaults: list[dict[str, str]],
    existing_rows: list[dict[str, str]],
    workbook_entries: list[dict[str, object]],
) -> list[dict[str, str]]:
    existing_by_key: dict[str, dict[str, str]] = {}
    existing_by_path: dict[str, list[dict[str, str]]] = {}
    for row in existing_rows:
        key = slug_or_id(row)
        if key:
            existing_by_key[key] = row
        path_key = normalize_windows_path(row.get("project_path"))
        if path_key:
            existing_by_path.setdefault(path_key, []).append(row)

    workbook_by_path = {entry["normalizedPath"]: entry for entry in workbook_entries}

    def match_existing_row(default: dict[str, str]) -> dict[str, str] | None:
        direct = existing_by_key.get(default["project_id"]) or existing_by_key.get(default["slug"])
        if direct:
            return direct

        candidates = existing_by_path.get(normalize_windows_path(default["project_path"]), [])
        if not candidates:
            return None
        if len(candidates) == 1:
            return candidates[0]

        name_key = normalize_name_key(default["project_name"])
        slug_key = normalize_name_key(default["slug"])
        base_id = tracker_base_id(default["project_id"])

        def score(row: dict[str, str]) -> tuple[int, str]:
            points = 0
            if normalize_name_key(row.get("project_name")) == name_key:
                points += 100
            if normalize_name_key(row.get("slug")) == slug_key:
                points += 80
            if tracker_base_id(row.get("project_id")) == base_id:
                points += 40
            return points, normalize_text(row.get("project_name"))

        best = max(candidates, key=score)
        return best if score(best)[0] > 0 else None

    merged: list[dict[str, str]] = []
    for default in defaults:
        existing = match_existing_row(default)
        row = dict(default)
        if existing:
            for column in MANUAL_COLUMNS:
                if column in existing:
                    row[column] = normalize_choice(column, existing[column]) if column in STATUS_CHOICES else normalize_text(existing[column])

        workbook_entry = workbook_by_path.get(normalize_windows_path(row["project_path"]))
        if workbook_entry:
            row["rewrite_workbook_source"] = r"C:\E156\rewrite-workbook.txt"
            row["rewrite_present"] = "yes" if workbook_entry["rewritePresent"] else "no"
            row["rewrite_valid"] = "yes" if workbook_entry["rewriteValid"] else "no"
            row["rewrite_word_count"] = str(workbook_entry["rewriteWordCount"]) if workbook_entry["rewritePresent"] else ""
            row["rewrite_sentence_count"] = str(workbook_entry["rewriteSentenceCount"]) if workbook_entry["rewritePresent"] else ""
            row["rewrite_matches_current"] = "yes" if workbook_entry["rewriteMatchesCurrent"] else "no"
            if workbook_entry["rewritePresent"]:
                row["rewrite_status"] = "rewritten" if workbook_entry["rewriteValid"] else "rewriting"
                row["publish_now"] = "yes"
                if workbook_entry["rewriteValid"]:
                    row["e156_body_status"] = "ready"
                if not row["source_article_path"]:
                    row["source_article_path"] = r"C:\E156\rewrite-workbook.txt"

        if row["protocol_status"] == "missing" and signal(row["source_protocol_signal"]):
            row["protocol_status"] = "draft"
        if row["dashboard_status"] == "missing" and row["existing_pages_url"]:
            row["dashboard_status"] = "draft"

        if row["project_id"].startswith("ext-"):
            metadata = load_submission_metadata(row["project_path"])
            path_state = inspect_project_path(row["project_path"])
            body_ready = signal(row["rewrite_valid"]) or metadata["bodyValid"] == "yes"
            body_draft = path_state["hasPaper"] or bool(metadata["body"])

            if not row["github_repo_url"]:
                row["github_repo_url"] = metadata["repoUrl"]
            if not row["github_pages_url"]:
                row["github_pages_url"] = metadata["siteUrl"]
            if not row["existing_pages_url"]:
                row["existing_pages_url"] = metadata["siteUrl"]

            row["github_status"] = "published" if row["github_repo_url"] else "missing"
            row["pages_status"] = "published" if row["github_pages_url"] else "missing"

            if row["e156_body_status"] != "published":
                row["e156_body_status"] = "ready" if body_ready else ("draft" if body_draft else "missing")

            source_article_path = row["source_article_path"]
            source_local = windows_to_local_path(source_article_path) if source_article_path and source_article_path != r"C:\E156\rewrite-workbook.txt" else None
            if not source_article_path or (source_local and not source_local.exists()):
                row["source_article_path"] = external_source_path(metadata, signal(row["rewrite_present"]))

        merged.append(row)

    for row in merged:
        for column in STATUS_CHOICES:
            row[column] = normalize_choice(column, row.get(column))

    merged.sort(
        key=lambda item: (
            PRIORITY_ORDER.get(item["portfolio_priority"], 99),
            item["project_name"].lower(),
        )
    )
    return merged


def write_tracker_csv(rows: list[dict[str, str]]) -> None:
    with TRACKER_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=TRACKER_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_tracker_json(rows: list[dict[str, str]], generated_at: str) -> None:
    payload = {
        "project": "SubmissionCockpit",
        "generatedAt": generated_at,
        "tracker": {
            "externalRewriteSource": r"C:\E156\rewrite-workbook.txt",
            "workbook": str(TRACKER_XLSX.relative_to(PROJECT_ROOT)),
            "csv": str(TRACKER_CSV.relative_to(PROJECT_ROOT)),
        },
        "columns": TRACKER_COLUMNS,
        "rows": rows,
    }
    write_json(TRACKER_JSON, payload)


def write_tracker_xlsx(rows: list[dict[str, str]]) -> None:
    if not OPENPYXL_AVAILABLE:
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tracker"
    sheet.append(TRACKER_COLUMNS)

    header_fill = PatternFill(fill_type="solid", fgColor="204436")
    header_font = Font(color="F9F5EE", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        sheet.append([row.get(column, "") for column in TRACKER_COLUMNS])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        "A": 12,
        "B": 28,
        "C": 24,
        "D": 28,
        "E": 12,
        "F": 18,
        "G": 26,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 16,
        "M": 12,
        "N": 12,
        "O": 28,
        "P": 20,
        "Q": 28,
        "R": 12,
        "S": 12,
        "T": 12,
        "U": 12,
        "V": 12,
        "W": 12,
        "X": 14,
        "Y": 12,
        "Z": 14,
        "AA": 14,
        "AB": 16,
        "AC": 16,
        "AD": 14,
        "AE": 12,
        "AF": 12,
        "AG": 18,
        "AH": 28,
        "AI": 28,
        "AJ": 28,
        "AK": 28,
        "AL": 36,
        "AM": 22,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    enums = workbook.create_sheet("Enums")
    enums["A1"] = "Column"
    enums["B1"] = "Allowed values"
    enums["A1"].font = Font(bold=True)
    enums["B1"].font = Font(bold=True)
    row_index = 2
    for column, choices in STATUS_CHOICES.items():
        enums[f"A{row_index}"] = column
        enums[f"B{row_index}"] = ", ".join(sorted(choices))
        row_index += 1

    notes = workbook.create_sheet("Notes")
    notes["A1"] = "How to use this workbook"
    notes["A1"].font = Font(bold=True)
    notes["A3"] = "Canonical rewrite source: C:\\E156\\rewrite-workbook.txt."
    notes["A4"] = "Edit the external text workbook for rewrites, then rerun scripts/build_submission_cockpit.py."
    notes["A5"] = "Edit the Tracker sheet for GitHub, Pages, PDF, and journal release states."
    notes["A6"] = "Canonical project metadata columns are regenerated from upstream portfolio snapshots."
    notes["A7"] = "Manual workflow columns are preserved and mirrored into CSV and JSON."
    notes["A8"] = "Default publication policy is E156 editorial + GitHub protocol + GitHub Pages + MIT code."
    notes.column_dimensions["A"].width = 120

    workbook.save(TRACKER_XLSX)


def signal(value: str) -> bool:
    return normalize_text(value).lower() in {"yes", "true", "1"}


def publication_score(record: dict[str, object]) -> int:
    score = 0
    if record["publishNow"]:
        score += 12
    score += min(int(record["opsReadinessScore"]) // 10, 10)
    score += min(int(record["citationReadinessScore"]) // 10, 10)
    score += min(int(record["workflowReadinessScore"]) // 20, 5)
    if record["rewriteStatus"] in REWRITE_READY_STATUSES:
        score += 20
    elif record["rewriteStatus"] == "rewriting":
        score += 10
    if record["e156BodyStatus"] in READY_STATUSES | PUBLISHED_STATUSES:
        score += 12
    elif record["e156BodyStatus"] == "draft":
        score += 6
    if record["protocolStatus"] in READY_STATUSES | PUBLISHED_STATUSES:
        score += 8
    elif record["protocolStatus"] == "draft":
        score += 4
    if record["dashboardStatus"] in READY_STATUSES | PUBLISHED_STATUSES:
        score += 8
    elif record["dashboardStatus"] == "draft":
        score += 4
    if record["codeReleaseStatus"] in READY_STATUSES | PUBLISHED_STATUSES:
        score += 8
    if record["mitLicenseStatus"] == "ready":
        score += 5
    elif record["mitLicenseStatus"] == "planned":
        score += 2
    if record["githubStatus"] in READY_STATUSES | PUBLISHED_STATUSES:
        score += 5
    if record["pagesStatus"] in READY_STATUSES | PUBLISHED_STATUSES:
        score += 4
    if record["pdfStatus"] in {"ready", "uploaded"}:
        score += 2
    if record["synthesisStatus"] in {"ready", "uploaded"}:
        score += 1
    return min(score, 100)


def blockers_for(record: dict[str, object]) -> list[str]:
    blockers: list[str] = []
    if record["publishNow"] and record["rewriteStatus"] not in REWRITE_READY_STATUSES:
        blockers.append("Finish rewrite and freeze E156 editorial text")
    if record["publishNow"] and record["e156BodyStatus"] not in READY_STATUSES | PUBLISHED_STATUSES:
        blockers.append("Mark the E156 paper bundle as ready")
    if record["publishNow"] and record["protocolStatus"] not in READY_STATUSES | PUBLISHED_STATUSES:
        blockers.append("Publish the protocol on GitHub")
    if record["publishNow"] and record["dashboardStatus"] not in READY_STATUSES | PUBLISHED_STATUSES:
        blockers.append("Build or confirm the HTML dashboard")
    if record["codeReleaseStatus"] not in READY_STATUSES | PUBLISHED_STATUSES:
        blockers.append("Prepare rerunnable code for release")
    if record["mitLicenseStatus"] != "ready":
        blockers.append("Assert MIT license in the release repo")
    if record["publishNow"] and record["githubStatus"] not in READY_STATUSES | PUBLISHED_STATUSES:
        blockers.append("Push the project repo to GitHub")
    if record["publishNow"] and record["pagesStatus"] not in READY_STATUSES | PUBLISHED_STATUSES:
        blockers.append("Publish or verify GitHub Pages")
    if record["publishNow"] and record["pdfStatus"] not in {"ready", "uploaded"}:
        blockers.append("Render the editorial PDF galley")
    if record["pdfStatus"] == "ready" and record["synthesisStatus"] not in {"ready", "uploaded"}:
        blockers.append("Queue the Synthesis Journal upload")
    if not blockers:
        blockers.append("Monitor and maintain")
    return blockers


def pipeline_stage(record: dict[str, object]) -> str:
    if record["publishNow"] is False:
        return "Backlog"
    if record["synthesisStatus"] == "uploaded":
        return "Uploaded"
    if (
        record["githubStatus"] in READY_STATUSES | PUBLISHED_STATUSES
        and record["pagesStatus"] in READY_STATUSES | PUBLISHED_STATUSES
        and record["e156BodyStatus"] in READY_STATUSES | PUBLISHED_STATUSES
        and record["protocolStatus"] in READY_STATUSES | PUBLISHED_STATUSES
        and record["dashboardStatus"] in READY_STATUSES | PUBLISHED_STATUSES
    ):
        return "Published"
    if record["rewriteStatus"] in REWRITE_READY_STATUSES:
        return "Packaging"
    if record["rewriteStatus"] == "rewriting":
        return "Rewriting"
    return "Queued"


def build_records(
    tracker_rows: list[dict[str, str]],
    portfolio: dict[str, object],
    ops: dict[str, object],
    citation: dict[str, object],
    authorship: dict[str, object],
    deposit: dict[str, object],
    generated_at: str,
) -> list[dict[str, object]]:
    portfolio_by_id = build_items_by_id(portfolio["portfolio"])
    ops_by_id = build_items_by_id(ops["projects"])
    citation_by_id = build_items_by_id(citation["projects"])
    authorship_by_id = build_items_by_id(authorship["projects"])
    deposit_by_id = build_items_by_id(deposit["items"])

    records: list[dict[str, object]] = []
    for row in tracker_rows:
        project_id = row["project_id"]
        raw_id = tracker_base_id(project_id)
        project = matched_item(
            portfolio_by_id,
            raw_id,
            project_name=row["project_name"],
            project_path=row["project_path"],
            project_slug=row["slug"],
        )
        ops_item = matched_item(
            ops_by_id,
            raw_id,
            project_name=row["project_name"],
            project_path=row["project_path"],
            project_slug=row["slug"],
        )
        citation_item = matched_item(
            citation_by_id,
            raw_id,
            project_name=row["project_name"],
            project_path=row["project_path"],
            project_slug=row["slug"],
        )
        authorship_item = matched_item(
            authorship_by_id,
            raw_id,
            project_name=row["project_name"],
            project_path=row["project_path"],
            project_slug=row["slug"],
        )
        deposit_item = matched_item(
            deposit_by_id,
            raw_id,
            project_name=row["project_name"],
            project_path=row["project_path"],
            project_slug=row["slug"],
        )
        is_indexed_project = bool(project)

        publish_now = normalize_choice("publish_now", row["publish_now"]) == "yes"

        record = {
            "id": project_id,
            "name": row["project_name"],
            "slug": row["slug"],
            "path": row["project_path"],
            "tier": row["tier"],
            "type": normalize_text(project.get("type")) or "methods",
            "resolvedStatus": row["resolved_status"],
            "targetJournal": row["target_journal"],
            "portfolioPriority": row["portfolio_priority"],
            "opsReadinessScore": int(row["ops_readiness_score"]),
            "citationReadinessScore": int(row["citation_readiness_score"]),
            "workflowReadinessScore": int(row["workflow_readiness_score"]),
            "recommendedLicense": row["recommended_license"],
            "sourcePaperSignal": signal(row["source_paper_signal"]),
            "sourceProtocolSignal": signal(row["source_protocol_signal"]),
            "existingPagesUrl": row["existing_pages_url"],
            "depositState": row["deposit_state"],
            "rewriteWorkbookSource": row["rewrite_workbook_source"],
            "rewritePresent": signal(row["rewrite_present"]),
            "rewriteValid": signal(row["rewrite_valid"]),
            "rewriteWordCount": int(row["rewrite_word_count"] or 0),
            "rewriteSentenceCount": int(row["rewrite_sentence_count"] or 0),
            "rewriteMatchesCurrent": signal(row["rewrite_matches_current"]),
            "rewriteStatus": normalize_choice("rewrite_status", row["rewrite_status"]),
            "publishNow": publish_now,
            "e156BodyStatus": normalize_choice("e156_body_status", row["e156_body_status"]),
            "protocolStatus": normalize_choice("protocol_status", row["protocol_status"]),
            "dashboardStatus": normalize_choice("dashboard_status", row["dashboard_status"]),
            "codeReleaseStatus": normalize_choice("code_release_status", row["code_release_status"]),
            "mitLicenseStatus": normalize_choice("mit_license_status", row["mit_license_status"]),
            "githubStatus": normalize_choice("github_status", row["github_status"]),
            "pagesStatus": normalize_choice("pages_status", row["pages_status"]),
            "pdfStatus": normalize_choice("pdf_status", row["pdf_status"]),
            "synthesisStatus": normalize_choice("synthesis_status", row["synthesis_status"]),
            "sourceArticlePath": row["source_article_path"],
            "githubRepoUrl": row["github_repo_url"],
            "githubPagesUrl": row["github_pages_url"],
            "paperPdfPath": row["paper_pdf_path"],
            "synthesisArticleUrl": row["synthesis_article_url"],
            "notes": row["notes"],
            "lastUpdated": row["last_updated"],
            "recordUrl": normalize_text(citation_item.get("recordUrl")) or row["existing_pages_url"],
            "packetUrl": normalize_text(authorship_item.get("packetUrl")),
            "cffUrl": normalize_text(authorship_item.get("cffUrl")),
            "depositDraftUrl": normalize_text(deposit_item.get("depositDraftUrl")),
            "queuePageUrl": normalize_text(deposit_item.get("queuePageUrl")),
            "primaryGap": normalize_text(authorship_item.get("primaryGap"))
            or normalize_text(citation_item.get("primaryGap"))
            or ("Not indexed in the portfolio snapshots yet" if not is_indexed_project else ""),
            "opsPrimaryAction": normalize_text(ops_item.get("primaryAction")) or ("Push GitHub + Pages release" if publish_now else ""),
            "generatedAt": generated_at,
        }
        record["publicationScore"] = publication_score(record)
        record["pipelineStage"] = pipeline_stage(record)
        record["blockers"] = blockers_for(record)
        record["readyForGitHubSync"] = (
            publish_now
            and record["rewriteStatus"] in REWRITE_READY_STATUSES
            and record["e156BodyStatus"] in READY_STATUSES | PUBLISHED_STATUSES
            and record["protocolStatus"] in READY_STATUSES | PUBLISHED_STATUSES
            and record["dashboardStatus"] in READY_STATUSES | PUBLISHED_STATUSES
            and record["codeReleaseStatus"] in READY_STATUSES | PUBLISHED_STATUSES
            and record["mitLicenseStatus"] == "ready"
            and record["githubStatus"] not in READY_STATUSES | PUBLISHED_STATUSES
            and not record["githubRepoUrl"]
        )
        record["readyForPages"] = (
            publish_now
            and record["githubStatus"] in READY_STATUSES | PUBLISHED_STATUSES
            and record["pagesStatus"] not in READY_STATUSES | PUBLISHED_STATUSES
            and not record["githubPagesUrl"]
        )
        record["readyForSynthesisUpload"] = (
            publish_now
            and record["pdfStatus"] == "ready"
            and record["synthesisStatus"] not in {"ready", "uploaded"}
        )
        records.append(record)

    records.sort(
        key=lambda item: (
            PRIORITY_ORDER.get(item["portfolioPriority"], 99),
            -item["publicationScore"],
            item["name"].lower(),
        )
    )
    return records


def build_dashboard_payload(
    records: list[dict[str, object]],
    generated_at: str,
    external_e156_status: dict[str, object],
) -> tuple[dict[str, object], dict[str, object]]:
    blocker_counts = Counter(record["blockers"][0] for record in records)
    rewrite_queue = [record for record in records if record["publishNow"] and record["rewriteStatus"] not in REWRITE_READY_STATUSES]
    build_queue = [
        record
        for record in records
        if record["publishNow"]
        and record["rewriteStatus"] in REWRITE_READY_STATUSES
        and not (
            record["e156BodyStatus"] in READY_STATUSES | PUBLISHED_STATUSES
            and record["protocolStatus"] in READY_STATUSES | PUBLISHED_STATUSES
            and record["dashboardStatus"] in READY_STATUSES | PUBLISHED_STATUSES
        )
    ]
    github_queue = [record for record in records if record["readyForGitHubSync"]]
    pages_queue = [record for record in records if record["readyForPages"]]
    synthesis_queue = [record for record in records if record["readyForSynthesisUpload"]]

    metrics = {
        "trackedProjects": len(records),
        "publishNowCount": sum(1 for record in records if record["publishNow"]),
        "rewrittenCount": sum(1 for record in records if record["rewriteStatus"] in REWRITE_READY_STATUSES),
        "workbookRewriteCount": sum(1 for record in records if record["rewritePresent"]),
        "workbookValidRewriteCount": sum(1 for record in records if record["rewriteValid"]),
        "bodyReadyCount": sum(1 for record in records if record["e156BodyStatus"] in READY_STATUSES | PUBLISHED_STATUSES),
        "protocolReadyCount": sum(1 for record in records if record["protocolStatus"] in READY_STATUSES | PUBLISHED_STATUSES),
        "dashboardReadyCount": sum(1 for record in records if record["dashboardStatus"] in READY_STATUSES | PUBLISHED_STATUSES),
        "githubReadyCount": sum(1 for record in records if record["githubStatus"] in READY_STATUSES | PUBLISHED_STATUSES),
        "pagesReadyCount": sum(1 for record in records if record["pagesStatus"] in READY_STATUSES | PUBLISHED_STATUSES),
        "readyForGitHubCount": len(github_queue),
        "readyForPagesCount": len(pages_queue),
        "readyForSynthesisCount": len(synthesis_queue),
        "citationHighCount": sum(1 for record in records if record["citationReadinessScore"] >= 90),
        "submissionReadyStatusCount": sum(1 for record in records if record["resolvedStatus"] == "Submission ready"),
    }

    cockpit_file = {
        "project": "SubmissionCockpit",
        "generatedAt": generated_at,
        "overview": metrics,
        "tracker": {
            "externalRewriteSource": r"C:\E156\rewrite-workbook.txt",
            "workbook": str(TRACKER_XLSX.relative_to(PROJECT_ROOT)),
            "csv": str(TRACKER_CSV.relative_to(PROJECT_ROOT)),
            "json": str(TRACKER_JSON.relative_to(PROJECT_ROOT)),
            "openpyxlAvailable": OPENPYXL_AVAILABLE,
        },
        "externalE156": external_e156_status,
        "projects": records,
    }

    dashboard = {
        "project": {
            "name": "SubmissionCockpit",
            "version": "0.1.0",
            "generatedAt": generated_at,
            "summary": "Canonical publication ledger for E156 editorials, driven by C:\\E156\\rewrite-workbook.txt and extended with GitHub, Pages, MIT, PDF, and Synthesis release states.",
            "externalRewriteSource": r"C:\E156\rewrite-workbook.txt",
            "trackerWorkbook": str(TRACKER_XLSX.relative_to(PROJECT_ROOT)),
            "trackerCsv": str(TRACKER_CSV.relative_to(PROJECT_ROOT)),
        },
        "metrics": metrics,
        "externalE156": external_e156_status,
        "topBlockers": [
            {"blocker": blocker, "count": count}
            for blocker, count in blocker_counts.most_common(8)
        ],
        "rewriteQueue": rewrite_queue[:20],
        "buildQueue": build_queue[:20],
        "releaseQueue": (github_queue + pages_queue)[:20],
        "synthesisQueue": synthesis_queue[:20],
        "projects": records,
    }
    return cockpit_file, dashboard


def write_exports(records: list[dict[str, object]], dashboard: dict[str, object]) -> None:
    for manifest_path in MANIFEST_DIR.glob("*.json"):
        manifest_path.unlink()

    for record in records:
        manifest = {
            "project": record["name"],
            "slug": record["slug"],
            "path": record["path"],
            "desiredRelease": {
                "paperType": "E156 editorial",
                "protocolHosting": "GitHub",
                "dashboardHosting": "GitHub Pages",
                "codeLicense": "MIT",
                "futureJournal": "Synthesis Journal",
            },
            "currentState": record,
        }
        write_json(MANIFEST_DIR / f"{record['slug']}.json", manifest)

    write_json(EXPORTS_DIR / "rewrite-queue.json", dashboard["rewriteQueue"])
    write_json(EXPORTS_DIR / "build-queue.json", dashboard["buildQueue"])
    write_json(EXPORTS_DIR / "github-queue.json", [record for record in records if record["readyForGitHubSync"]])
    write_json(EXPORTS_DIR / "pages-queue.json", [record for record in records if record["readyForPages"]])
    write_json(EXPORTS_DIR / "synthesis-queue.json", dashboard["synthesisQueue"])


def build_e156_index() -> None:
    if not E156_BUILDER.exists() or not E156_PAPER_JSON.exists():
        return
    subprocess.run(
        [sys.executable, str(E156_BUILDER), "--input", str(E156_PAPER_JSON), "--output", str(E156_INDEX_HTML)],
        check=True,
    )


def main() -> None:
    ensure_dirs()
    sync_snapshots()
    generated_at = now_iso()

    portfolio = load_json(SNAPSHOT_FILES["portfolio"])
    ops = load_json(SNAPSHOT_FILES["ops"])
    citation = load_json(SNAPSHOT_FILES["citation"])
    authorship = load_json(SNAPSHOT_FILES["authorship"])
    deposit = load_json(SNAPSHOT_FILES["deposit"])
    external_workbook_entries = parse_external_rewrite_workbook(SNAPSHOT_FILES["external_rewrite_workbook"])
    external_e156_status = build_external_e156_status(external_workbook_entries)

    portfolio_by_id = build_items_by_id(portfolio["portfolio"])
    ops_by_id = build_items_by_id(ops["projects"])
    citation_by_id = build_items_by_id(citation["projects"])
    authorship_by_id = build_items_by_id(authorship["projects"])
    deposit_by_id = build_items_by_id(deposit["items"])
    duplicate_counts = {project_id: len(items) for project_id, items in portfolio_by_id.items()}

    defaults = [
        default_tracker_row(
            unique_indexed_project_id(
                project,
                matched_item(
                    citation_by_id,
                    normalize_text(project["id"]),
                    project_name=project.get("name"),
                    project_path=project.get("path"),
                ),
                duplicate_counts,
            ),
            project,
            matched_item(
                citation_by_id,
                normalize_text(project["id"]),
                project_name=project.get("name"),
                project_path=project.get("path"),
            ),
            matched_item(
                authorship_by_id,
                normalize_text(project["id"]),
                project_name=project.get("name"),
                project_path=project.get("path"),
            ),
            matched_item(
                deposit_by_id,
                normalize_text(project["id"]),
                project_name=project.get("name"),
                project_path=project.get("path"),
            ),
            matched_item(
                ops_by_id,
                normalize_text(project["id"]),
                project_name=project.get("name"),
                project_path=project.get("path"),
            ),
        )
        for project in portfolio["portfolio"]
    ]
    default_paths = {normalize_windows_path(row["project_path"]) for row in defaults}
    repo_mapping = load_repo_mapping()
    defaults.extend(
        [
            default_external_tracker_row(entry, repo_mapping)
            for entry in external_workbook_entries
            if entry["normalizedPath"] not in default_paths
        ]
    )

    existing_tracker_rows = load_tracker_xlsx() or load_tracker_csv()
    tracker_rows = merge_tracker_rows(defaults, existing_tracker_rows, external_workbook_entries)
    write_tracker_csv(tracker_rows)
    write_tracker_json(tracker_rows, generated_at)
    write_tracker_xlsx(tracker_rows)

    records = build_records(tracker_rows, portfolio, ops, citation, authorship, deposit, generated_at)
    cockpit_file, dashboard = build_dashboard_payload(records, generated_at, external_e156_status)

    write_json(COCKPIT_JSON, cockpit_file)
    write_json(DATA_JSON, dashboard)
    DATA_JS.write_text("window.SUBMISSION_COCKPIT_DATA = " + json.dumps(dashboard, indent=2) + ";\n", encoding="utf-8")
    write_exports(records, dashboard)
    build_e156_index()

    print(
        "Built SubmissionCockpit "
        f"({dashboard['metrics']['trackedProjects']} projects, "
        f"{dashboard['metrics']['publishNowCount']} marked publish-now, "
        f"{dashboard['metrics']['readyForGitHubCount']} ready for GitHub sync)."
    )


if __name__ == "__main__":
    main()
